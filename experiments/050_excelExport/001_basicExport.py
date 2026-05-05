"""
실험 ID: 001
실험명: openpyxl 기본 Excel 내보내기

목적:
- dartlab Company 데이터를 openpyxl로 .xlsx 파일 생성
- 손익계산서(IS), 재무상태표(BS), 현금흐름표(CF)를 각각 시트로 분리
- 연도별 시계열을 가로축으로, 계정을 세로축으로 배치
- 기본 서식: 헤더 볼드, 숫자 천단위 콤마, 열 너비 자동조절

가설:
1. buildAnnual()의 연도별 시계열을 직접 Excel로 변환하면 가독성 확보 가능
2. openpyxl의 서식 기능으로 최소한의 가공만으로 실무 품질 달성 가능

방법:
1. Company("005930") → buildAnnual()로 연도별 series + years 획득
2. IS/BS/CF 각각 시트 생성
3. 1행: 헤더(계정명, 연도...), 2행~: 계정별 시계열
4. 서식: 헤더 bold+배경, 숫자 #,##0, 열 너비 자동

결과 (실험 후 작성):
-

결론:
-

실험일: 2026-03-10
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dartlab.providers.dart.finance.pivot import buildAnnual

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

ACCOUNT_FONT = Font(bold=True, size=10)
NUMBER_FORMAT = '#,##0'
NEGATIVE_FORMAT = '#,##0;[Red]-#,##0'

SJ_LABELS = {
	"IS": "손익계산서",
	"BS": "재무상태표",
	"CF": "현금흐름표",
}

DISPLAY_ORDER = {
	"IS": [
		"revenue", "cost_of_sales", "gross_profit",
		"selling_and_administrative_expenses", "operating_income",
		"other_income", "other_expenses",
		"finance_income", "finance_cost",
		"profit_before_tax", "income_tax_expense", "net_income",
		"comprehensive_income",
	],
	"BS": [
		"current_assets", "cash_and_equivalents", "shortterm_financial_instruments",
		"trade_receivables", "inventories",
		"non_current_assets", "ppe", "intangible_assets",
		"total_assets",
		"current_liabilities", "non_current_liabilities", "total_liabilities",
		"total_equity", "equity_nci",
		"retained_earnings", "issued_capital",
	],
	"CF": [
		"operating_cashflow", "investing_cashflow", "financing_cashflow",
	],
}

ACCOUNT_LABELS = {
	"revenue": "매출액",
	"cost_of_sales": "매출원가",
	"gross_profit": "매출총이익",
	"selling_and_administrative_expenses": "판관비",
	"operating_income": "영업이익",
	"other_income": "기타수익",
	"other_expenses": "기타비용",
	"finance_income": "금융수익",
	"finance_cost": "금융비용",
	"profit_before_tax": "법인세차감전이익",
	"income_tax_expense": "법인세비용",
	"net_income": "당기순이익",
	"comprehensive_income": "총포괄이익",
	"current_assets": "유동자산",
	"cash_and_equivalents": "현금및현금성자산",
	"shortterm_financial_instruments": "단기금융상품",
	"trade_receivables": "매출채권",
	"inventories": "재고자산",
	"non_current_assets": "비유동자산",
	"ppe": "유형자산",
	"intangible_assets": "무형자산",
	"total_assets": "자산총계",
	"current_liabilities": "유동부채",
	"non_current_liabilities": "비유동부채",
	"total_liabilities": "부채총계",
	"total_equity": "자본총계(지배)",
	"equity_nci": "비지배지분",
	"retained_earnings": "이익잉여금",
	"issued_capital": "자본금",
	"operating_cashflow": "영업활동CF",
	"investing_cashflow": "투자활동CF",
	"financing_cashflow": "재무활동CF",
}


def _autoWidth(ws, minWidth=10, maxWidth=25):
	for col in ws.columns:
		maxLen = minWidth
		colLetter = get_column_letter(col[0].column)
		for cell in col:
			if cell.value is not None:
				cellLen = len(str(cell.value))
				if isinstance(cell.value, (int, float)):
					cellLen = max(cellLen, 12)
				if cellLen > maxLen:
					maxLen = cellLen
		ws.column_dimensions[colLetter].width = min(maxLen + 2, maxWidth)


def _writeSheet(wb, sjDiv, series, years):
	label = SJ_LABELS.get(sjDiv, sjDiv)
	ws = wb.create_sheet(title=label)

	headers = ["계정"] + years
	for colIdx, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=colIdx, value=header)
		cell.font = HEADER_FONT
		cell.fill = HEADER_FILL
		cell.alignment = HEADER_ALIGN

	stmtData = series.get(sjDiv, {})
	displayKeys = DISPLAY_ORDER.get(sjDiv, list(stmtData.keys()))

	row = 2
	for snakeId in displayKeys:
		vals = stmtData.get(snakeId)
		if vals is None:
			continue

		accountLabel = ACCOUNT_LABELS.get(snakeId, snakeId)
		cell = ws.cell(row=row, column=1, value=accountLabel)
		cell.font = ACCOUNT_FONT

		for colIdx, val in enumerate(vals, 2):
			if val is not None:
				cell = ws.cell(row=row, column=colIdx, value=round(val))
				cell.number_format = NEGATIVE_FORMAT

		row += 1

	otherKeys = [k for k in stmtData if k not in displayKeys and k != "periods"]
	if otherKeys:
		row += 1
		sep = ws.cell(row=row, column=1, value="── 기타 계정 ──")
		sep.font = Font(italic=True, color="888888")
		row += 1

		for snakeId in sorted(otherKeys):
			vals = stmtData[snakeId]
			allNone = all(v is None for v in vals)
			if allNone:
				continue

			accountLabel = ACCOUNT_LABELS.get(snakeId, snakeId)
			cell = ws.cell(row=row, column=1, value=accountLabel)

			for colIdx, val in enumerate(vals, 2):
				if val is not None:
					cell = ws.cell(row=row, column=colIdx, value=round(val))
					cell.number_format = NEGATIVE_FORMAT

			row += 1

	_autoWidth(ws)

	ws.freeze_panes = "B2"


def exportBasic(stockCode, outputPath=None):
	result = buildAnnual(stockCode)
	if result is None:
		print(f"[ERROR] {stockCode}: 재무 데이터 없음")
		return None

	series, years = result

	wb = Workbook()
	wb.remove(wb.active)

	for sjDiv in ["IS", "BS", "CF"]:
		if sjDiv in series:
			_writeSheet(wb, sjDiv, series, years)

	if outputPath is None:
		outputPath = Path(__file__).parent / f"{stockCode}_financial.xlsx"

	wb.save(str(outputPath))
	print(f"[OK] {outputPath}")
	print(f"     시트: {wb.sheetnames}")
	print(f"     연도: {years}")

	for sjDiv in ["IS", "BS", "CF"]:
		stmtData = series.get(sjDiv, {})
		total = len(stmtData)
		nonEmpty = sum(1 for v in stmtData.values() if any(x is not None for x in v))
		print(f"     {sjDiv}: {nonEmpty}/{total} 계정에 데이터")

	return str(outputPath)


if __name__ == "__main__":
	exportBasic("005930")
