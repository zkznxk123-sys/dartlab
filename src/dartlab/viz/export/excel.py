"""Company → Excel (.xlsx) 내보내기.

Company의 모든 property를 동적으로 스캔하여 DataFrame이면 시트로 생성.
finance 시계열(IS/BS/CF)은 연도별 피벗, 나머지는 원본 DataFrame 그대로.

사용법::

    from dartlab import Company
    c = Company("005930")
    c.toExcel()                          # 005930_삼성전자.xlsx (전체)
    c.toExcel("output.xlsx")             # 지정 경로
    c.toExcel(modules=["IS", "BS"])      # 시트 선택
    c.toExcel(modules=["dividend", "audit", "employee"])  # 원하는 모듈만

    # 템플릿 기반 내보내기
    from dartlab.viz.export.template import PRESETS
    exportWithTemplate(c, PRESETS["summary"])
    exportWithTemplate(c, PRESETS["governance"], "output.xlsx")
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Optional

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dartlab.core.financeDocAccessor import getFinanceDocAccessor
from dartlab.synth.ratioCategories import RATIO_CATEGORIES

if TYPE_CHECKING:
    from dartlab.core.protocols import CompanyProtocol as Company
    from dartlab.providers.dart.parse.htmlTableParser import HtmlTableCell
    from dartlab.viz.export.template import ExcelTemplate, PanelTableSource


_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_ACCOUNT_FONT = Font(bold=True, size=10)
_NEGATIVE_FMT = "#,##0;[Red]-#,##0"
_NUMBER_FMT = "#,##0"

_SJ_LABELS = {"IS": "손익계산서", "BS": "재무상태표", "CF": "현금흐름표", "RATIO": "재무비율"}

_ACCOUNT_LABELS_OVERRIDE: dict[str, str] = {
    "owners_of_parent_equity": "자본총계(지배)",
    "operating_cashflow": "영업활동CF",
    "investing_cashflow": "투자활동CF",
    "cash_flows_from_financing_activities": "재무활동CF",
    "capex": "유형자산취득",
    "rnd_expenses": "연구개발비",
    "amortization": "무형자산상각비",
}

_RATIO_LABELS: dict[str, str] = {
    "roe": "ROE (%)",
    "roa": "ROA (%)",
    "operatingMargin": "영업이익률 (%)",
    "netMargin": "순이익률 (%)",
    "grossMargin": "매출총이익률 (%)",
    "ebitdaMargin": "EBITDA마진 (%)",
    "costOfSalesRatio": "매출원가율 (%)",
    "sgaRatio": "판관비율 (%)",
    "debtRatio": "부채비율 (%)",
    "currentRatio": "유동비율 (%)",
    "quickRatio": "당좌비율 (%)",
    "equityRatio": "자기자본비율 (%)",
    "interestCoverage": "이자보상배율 (x)",
    "netDebtRatio": "순차입금비율 (%)",
    "noncurrentRatio": "비유동비율 (%)",
    "revenueGrowth": "매출성장률 (%)",
    "operatingProfitGrowth": "영업이익성장률 (%)",
    "netProfitGrowth": "순이익성장률 (%)",
    "assetGrowth": "자산성장률 (%)",
    "equityGrowthRate": "자본성장률 (%)",
    "totalAssetTurnover": "총자산회전율 (x)",
    "inventoryTurnover": "재고자산회전율 (x)",
    "receivablesTurnover": "매출채권회전율 (x)",
    "payablesTurnover": "매입채무회전율 (x)",
    "fcf": "FCF",
    "operatingCfMargin": "영업CF마진 (%)",
    "operatingCfToNetIncome": "영업CF/순이익 (%)",
    "capexRatio": "CAPEX비율 (%)",
    "dividendPayoutRatio": "배당성향 (%)",
    "revenue": "매출",
    "operatingProfit": "영업이익",
    "netProfit": "순이익",
    "totalAssets": "총자산",
    "totalEquity": "자본총계",
    "operatingCashflow": "영업CF",
}

_CATEGORY_LABELS: dict[str, str] = {
    "profitability": "수익성",
    "stability": "안정성",
    "growth": "성장성",
    "efficiency": "효율성",
    "cashflow": "현금흐름",
    "absolute": "절대지표",
}


def _buildAccountLabels() -> dict[str, str]:
    """mapper.labelMap() 기반 + override 합성."""
    accessor = getFinanceDocAccessor()
    labels = accessor.accountLabels() if accessor else {}
    labels.update(_ACCOUNT_LABELS_OVERRIDE)
    return labels


_ACCOUNT_LABELS: dict[str, str] = {}


def _getAccountLabels() -> dict[str, str]:
    """lazy 초기화."""
    global _ACCOUNT_LABELS
    if not _ACCOUNT_LABELS:
        _ACCOUNT_LABELS = _buildAccountLabels()
    return _ACCOUNT_LABELS


_FINANCE_SHEETS = frozenset({"IS", "BS", "CF"})
_SPECIAL_SHEETS = frozenset({"ratios"})


def _hasFlag(c: Company, name: str) -> bool:
    return getattr(c, name, False) is True


def _hasFinance(c: Company) -> bool:
    return _hasFlag(c, "_hasFinance") or _hasFlag(c, "_hasFinanceParquet")


def _buildAnnualSeries(c: Company):
    accessor = getFinanceDocAccessor()
    if accessor is not None:
        result = accessor.buildAnnual(c.stockCode)
        if result:
            return result
    builder = getattr(c, "_buildFinanceSeries", None)
    if callable(builder):
        result = builder(freq="Y")
        if result:
            return result
    return None


def _autoWidth(ws, minWidth: int = 12, maxWidth: int = 30) -> None:
    for col in ws.columns:
        colLetter = get_column_letter(col[0].column)
        best = minWidth
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                if isinstance(cell.value, (int, float)):
                    length = max(length, 14)
                if length > best:
                    best = length
        ws.column_dimensions[colLetter].width = min(best + 2, maxWidth)


def _writeHeaders(ws, headers: list[str]) -> None:
    for colIdx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=colIdx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _writeFinanceSheet(
    wb: Workbook,
    sjDiv: str,
    series: dict[str, dict[str, list[Optional[float]]]],
    years: list[str],
    *,
    label: str = "",
    columns: list[str] | None = None,
    yearFilter: list[str] | None = None,
) -> None:
    sheetLabel = label or _SJ_LABELS.get(sjDiv, sjDiv)
    ws = wb.create_sheet(title=sheetLabel[:31])

    stmtData = series.get(sjDiv, {})
    if not stmtData:
        return

    displayYears = years
    yearIndices: list[int] | None = None
    if yearFilter:
        yearIndices = [i for i, y in enumerate(years) if y in yearFilter]
        displayYears = [years[i] for i in yearIndices]

    _writeHeaders(ws, ["계정"] + displayYears)

    row = 2
    for snakeId, vals in stmtData.items():
        if columns and snakeId not in columns:
            continue
        effectiveVals = [vals[i] for i in yearIndices] if yearIndices else vals
        if not any(v is not None for v in effectiveVals):
            continue
        ws.cell(row=row, column=1, value=_getAccountLabels().get(snakeId, snakeId)).font = _ACCOUNT_FONT
        for colIdx, val in enumerate(effectiveVals, 2):
            if val is not None:
                cell = ws.cell(row=row, column=colIdx, value=round(val))
                cell.number_format = _NEGATIVE_FMT
        row += 1

    _autoWidth(ws)
    ws.freeze_panes = "B2"


def _writeRatiosSheet(wb: Workbook, c: Company, *, label: str = "") -> None:
    from dartlab.analysis.financial.ratios import calcRatioSeries, toSeriesDict

    annualResult = c._buildFinanceSeries(freq="Y")
    if annualResult is None or len(annualResult) != 2:
        return
    annualSeries, years = annualResult
    if not annualSeries or not years:
        return
    rs = calcRatioSeries(annualSeries, years)
    ratioSeriesDict, _ = toSeriesDict(rs)
    ratioData = ratioSeriesDict.get("RATIO", {})
    if not ratioData:
        return

    ws = wb.create_sheet(title=(label or "재무비율")[:31])
    _writeHeaders(ws, ["지표"] + years)

    _CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    _CATEGORY_FONT = Font(bold=True, size=11, color="1F3864")

    _ABS_FIELDS = {"fcf", "revenue", "operatingProfit", "netProfit", "totalAssets", "totalEquity", "operatingCashflow"}

    row = 2
    for catKey, fields in RATIO_CATEGORIES:
        catFields = [f for f in fields if f in ratioData]
        if not catFields:
            continue

        catLabel = _CATEGORY_LABELS.get(catKey, catKey)
        cell = ws.cell(row=row, column=1, value=catLabel)
        cell.font = _CATEGORY_FONT
        cell.fill = _CATEGORY_FILL
        for colIdx in range(2, len(years) + 2):
            ws.cell(row=row, column=colIdx).fill = _CATEGORY_FILL
        row += 1

        for fieldName in catFields:
            vals = ratioData[fieldName]
            displayLabel = _RATIO_LABELS.get(fieldName, fieldName)
            ws.cell(row=row, column=1, value=displayLabel).font = _ACCOUNT_FONT
            isAbs = fieldName in _ABS_FIELDS
            for colIdx, val in enumerate(vals, 2):
                if val is not None:
                    cell = ws.cell(row=row, column=colIdx, value=round(val) if isAbs else round(val, 2))
                    cell.number_format = _NEGATIVE_FMT if isAbs else "0.00"
            row += 1

    _autoWidth(ws)
    ws.freeze_panes = "B2"


def _writeDataFrameSheet(
    wb: Workbook,
    title: str,
    df: pl.DataFrame,
    *,
    columns: list[str] | None = None,
    maxRows: int = 0,
) -> None:
    if df.height == 0:
        return

    if columns:
        validCols = [c for c in columns if c in df.columns]
        if validCols:
            df = df.select(validCols)

    if maxRows > 0:
        df = df.head(maxRows)

    safeTitle = title[:31]
    existing = {ws.title for ws in wb.worksheets}
    if safeTitle in existing:
        safeTitle = safeTitle[:28] + "_2"

    ws = wb.create_sheet(title=safeTitle)
    _writeHeaders(ws, df.columns)

    for rowIdx, row in enumerate(df.iter_rows(named=True), 2):
        for colIdx, colName in enumerate(df.columns, 1):
            val = row[colName]
            if val is None:
                continue
            cell = ws.cell(row=rowIdx, column=colIdx, value=val)
            if isinstance(val, float):
                if abs(val) >= 1000:
                    cell.number_format = _NEGATIVE_FMT
                else:
                    cell.number_format = "0.00"
            elif isinstance(val, int):
                cell.number_format = _NUMBER_FMT

    _autoWidth(ws)
    ws.freeze_panes = "A2"


def _getAvailableModules(c: Company) -> list[tuple[str, str]]:
    from dartlab.core.registry import getEntries

    accessor = getFinanceDocAccessor()
    available = [("IS", "손익계산서"), ("BS", "재무상태표"), ("CF", "현금흐름표")]
    known = {name for name, _ in available}
    for category in ("report", "disclosure"):
        for entry in getEntries(category=category):
            if entry.name not in known:
                available.append((entry.name, entry.label))
                known.add(entry.name)
    if accessor:
        try:
            accessorModules = accessor.exportModules()
        except (RuntimeError, TypeError, ValueError):
            accessorModules = []
        for name, label in accessorModules:
            if name not in known:
                available.append((name, label))
                known.add(name)
    if "ratios" not in known:
        available.append(("ratios", "재무비율"))
    return available


def _resolveData(c: Company, moduleName: str) -> pl.DataFrame | None:
    if moduleName == "ratios":
        return None
    try:
        data = c.panel(moduleName)
    except (KeyError, ValueError, TypeError):
        data = None
    if isinstance(data, pl.DataFrame) and data.height > 0:
        return data
    show = getattr(c, "show", None)
    if callable(show):
        try:
            data = show(moduleName)
        except (KeyError, ValueError, TypeError):
            data = None
    if isinstance(data, pl.DataFrame) and data.height > 0:
        return data
    data = getattr(c, moduleName, None)
    if isinstance(data, pl.DataFrame) and data.height > 0:
        return data
    return None


# ── 공시 panel 표 (병합 보존 격자) export ──

# panel wide 행을 유일 식별하는 7-필드 (실측: 005930 1627행 0충돌). disclosureKey/scope/leafSeq 는
# 대부분의 공시 표에서 None 이라 선택 — 제공된(non-None) 필드만으로 필터한다.
_PANEL_ID_FIELDS = ("chapter", "sectionLeaf", "blockLeaf", "leafType", "disclosureKey", "scope", "leafSeq")

_NOTE_FMT = Font(italic=True, size=9, color="808080")


def _loadPanelWide(c: Company) -> "pl.DataFrame | None":
    """회사 panel wide DataFrame (raw XML 셀) 로드 — c.panel 우선, 없으면 readWide.

    Args:
        c: Company 인스턴스 (stockCode 필요).

    Returns:
        wide pl.DataFrame (index 컬럼 + period 컬럼, 셀=raw XML) 또는 None.

    Example:
        >>> w = _loadPanelWide(c)  # doctest: +SKIP

    Raises:
        없음 — 로드 실패는 None.
    """
    from dartlab.providers.dart.panel.read import readWide

    marketNs = getattr(c, "marketNs", "kr") or "kr"
    try:
        return readWide(c.stockCode, marketNs=marketNs, tag=True)
    except (RuntimeError, ValueError, OSError):
        return None


def _selectPanelRow(wide: "pl.DataFrame", spec: PanelTableSource) -> dict | None:
    """panel wide 에서 spec 이 가리키는 단일 표 행 1개 선택 (제공 필드만 매칭 + leafSeq 디스앰비그).

    spec 의 non-None 식별 필드(_PANEL_ID_FIELDS)로 필터한다. disclosureKey/scope/leafSeq 가
    None 이면 그 필드는 매칭에서 제외(best-effort). 여러 행이 남으면 leafSeq 오름차순 첫 행.

    Args:
        wide: panel wide DataFrame (index 컬럼 + period 컬럼).
        spec: PanelTableSource (식별 필드 보유).

    Returns:
        선택된 행 dict (named) 또는 매칭 0 이면 None.

    Example:
        >>> row = _selectPanelRow(wide, spec)  # doctest: +SKIP

    Raises:
        없음 — 매칭 실패는 None.
    """
    flt = wide
    for fld in _PANEL_ID_FIELDS:
        val = getattr(spec, fld, None)
        if val is None or fld not in wide.columns:
            continue
        flt = flt.filter(pl.col(fld) == val)
    if flt.height == 0:
        return None
    if flt.height > 1 and "leafSeq" in flt.columns:
        flt = flt.sort("leafSeq")
    return flt.head(1).to_dicts()[0]


def _mergeRanges(grid: "list[list[HtmlTableCell]]") -> "dict[int, tuple[int, int, int, int]]":
    """격자에서 병합 범위 추출 — 셀 인스턴스 id() → (minRow, minCol, maxRow, maxCol).

    병합셀은 cellGrid 가 같은 HtmlTableCell 인스턴스를 여러 좌표에 둔다. id() 로 dedup 해 각
    병합의 경계(extent)를 구한다 — 극단 rowspan(683) 도 단일 범위 1개로 (N회 쓰기 아님).

    Args:
        grid: cellGrid 결과 (병합셀 = 동일 인스턴스 공유).

    Returns:
        ``{id(cell): (minR, minC, maxR, maxC)}`` — 모든 셀(병합·단일 공통).

    Example:
        >>> ext = _mergeRanges(grid)  # doctest: +SKIP

    Raises:
        없음.
    """
    coords: dict[int, list[tuple[int, int]]] = {}
    for r, row in enumerate(grid):
        for cIdx, cell in enumerate(row):
            if cell is None:
                continue
            coords.setdefault(id(cell), []).append((r, cIdx))
    ext: dict[int, tuple[int, int, int, int]] = {}
    for cid, cl in coords.items():
        rs = [r for r, _ in cl]
        cs = [c for _, c in cl]
        ext[cid] = (min(rs), min(cs), max(rs), max(cs))
    return ext


def _writeGridSheet(
    wb: Workbook,
    label: str,
    grid: "list[list[HtmlTableCell]]",
    *,
    unit: str = "",
    note: str = "",
) -> None:
    """병합 보존 격자를 openpyxl 시트로 — 앵커 1회 쓰기 + merge_cells, 값=coerceCell.

    Args:
        wb: openpyxl Workbook.
        label: 시트 라벨 (31자 trim, 충돌 회피).
        grid: cellGrid 결과 (병합셀 = 동일 인스턴스).
        unit: 단위 토큰 (detectUnit, 캡션 텍스트에서 — 셀 아님). 비면 단위 행 생략.
        note: 시트 노트 (예 "수평화 미지원(원본 구조)") — 비면 생략. 값 환산 0.

    Returns:
        None — wb 에 시트 추가 (부작용).

    Example:
        >>> _writeGridSheet(wb, "개요표", grid, unit="백만원")  # doctest: +SKIP

    Raises:
        없음 — 빈 격자는 빈 시트 생성 안 함(early return).
    """
    from dartlab.providers.dart.parse.dartXmlNormalize import coerceCell

    if not grid:
        return
    safeTitle = (label or "표")[:31]
    existing = {ws.title for ws in wb.worksheets}
    if safeTitle in existing:
        safeTitle = (safeTitle[:28] + "_2")[:31]
    ws = wb.create_sheet(title=safeTitle)

    startRow = 1  # openpyxl 1-base
    # 단위·노트 머리 행 (값 환산 없음 — 라벨만).
    if unit:
        cell = ws.cell(row=startRow, column=1, value=f"(단위: {unit})")
        cell.font = _NOTE_FMT
        startRow += 1
    if note:
        cell = ws.cell(row=startRow, column=1, value=note)
        cell.font = _NOTE_FMT
        startRow += 1

    ext = _mergeRanges(grid)
    written: set[int] = set()
    nCols = max((len(r) for r in grid), default=0)
    for r, row in enumerate(grid):
        for cIdx in range(nCols):
            cell = row[cIdx] if cIdx < len(row) else None
            if cell is None:
                continue
            cid = id(cell)
            if cid in written:
                continue  # 병합셀 — 앵커에만 1회 쓰기
            written.add(cid)
            minR, minC, maxR, maxC = ext[cid]
            value = coerceCell(cell.text)
            xlRow = startRow + minR
            xlCol = minC + 1
            oc = ws.cell(row=xlRow, column=xlCol, value=value)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                oc.number_format = _NEGATIVE_FMT
            if cell.isHeader:
                oc.font = _ACCOUNT_FONT
            align = "right" if (cell.align == "right" or isinstance(value, (int, float))) else (cell.align or "left")
            wrap = "\n" in cell.text if isinstance(cell.text, str) else False
            oc.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            # 극단 rowspan(683) 포함 — 단일 merge 범위 1개.
            if maxR > minR or maxC > minC:
                ws.merge_cells(
                    start_row=startRow + minR,
                    start_column=minC + 1,
                    end_row=startRow + maxR,
                    end_column=maxC + 1,
                )
    _autoWidth(ws)


def _writePanelTableSheet(wb: Workbook, spec: PanelTableSource, c: Company, *, label: str) -> None:
    """PanelTableSource → 시트. asFiled=단일 기간 격자, horizontalized=항목×기간/폴백.

    Args:
        wb: openpyxl Workbook.
        spec: PanelTableSource (식별 + periodMode + period).
        c: Company 인스턴스.
        label: 시트 라벨.

    Returns:
        None — wb 에 시트 추가. 행/기간 부재면 무동작(회귀 0).

    Example:
        >>> _writePanelTableSheet(wb, spec, c, label="개요표")  # doctest: +SKIP

    Raises:
        없음 — panel/행/기간 부재는 graceful skip.
    """
    from dartlab.providers.dart.parse.dartXmlNormalize import detectUnit, normalizeDartXml
    from dartlab.providers.dart.parse.htmlTableParser import cellGrid

    wide = _loadPanelWide(c)
    if wide is None:
        return
    row = _selectPanelRow(wide, spec)
    if row is None:
        return

    if spec.periodMode == "horizontalized":
        _writeHorizontalizedSheet(wb, label, wide, row, spec)
        return

    # asFiled — 단일 기간 원본 격자.
    period = spec.period
    if not period or period not in row or not row[period]:
        # 기간 미지정/부재 → 셀이 있는 최신 기간 자동 선택.
        idxCols = set(_PANEL_ID_FIELDS) | {"leafSeq"}
        periodCols = [c2 for c2 in wide.columns if c2 not in idxCols]
        period = next((p for p in periodCols if row.get(p)), None)
    if not period or not row.get(period):
        return
    rawXml = row[period]
    grid = cellGrid(normalizeDartXml(rawXml))
    unit = detectUnit(rawXml)
    _writeGridSheet(wb, label, grid, unit=unit)


def _writeHorizontalizedSheet(
    wb: Workbook,
    label: str,
    wide: "pl.DataFrame",
    row: dict,
    spec: PanelTableSource,
) -> None:
    """horizontalized 표 — 일반 공시 표는 라벨 정렬 불확실 → as-filed 폴백 + 노트 (honest-gap).

    일반 공시 표는 canonical account_id 가 없어 기간 간 행 라벨 정렬이 불확실하다. 거짓 정렬을
    만들지 않고 최신 기간의 as-filed 격자를 쓰되 "수평화 미지원(원본 구조)" 노트를 단다. finance
    정량(IS/BS/CF)은 ModuleSource 경로(_writeFinanceSheet)가 이미 항목×기간을 다룬다.

    Args:
        wb: openpyxl Workbook.
        label: 시트 라벨.
        wide: panel wide DataFrame.
        row: _selectPanelRow 결과 행.
        spec: PanelTableSource.

    Returns:
        None — wb 에 시트 추가 (as-filed 격자 + 노트).

    Example:
        >>> _writeHorizontalizedSheet(wb, "표", wide, row, spec)  # doctest: +SKIP

    Raises:
        없음.
    """
    from dartlab.providers.dart.parse.dartXmlNormalize import detectUnit, normalizeDartXml
    from dartlab.providers.dart.parse.htmlTableParser import cellGrid

    idxCols = set(_PANEL_ID_FIELDS) | {"leafSeq"}
    periodCols = [c2 for c2 in wide.columns if c2 not in idxCols]
    period = next((p for p in periodCols if row.get(p)), None)
    if not period or not row.get(period):
        return
    rawXml = row[period]
    grid = cellGrid(normalizeDartXml(rawXml))
    unit = detectUnit(rawXml)
    _writeGridSheet(
        wb,
        label,
        grid,
        unit=unit,
        note=f"수평화 미지원(원본 구조) — as-filed {period}",
    )


def exportToExcel(
    c: Company,
    outputPath: str | Path | None = None,
    modules: list[str] | None = None,
) -> str:
    """Company 데이터를 .xlsx로 내보낸다.

    Args:
            c: Company 인스턴스.
            outputPath: 저장 경로. None이면 현재 디렉토리에 '{종목코드}_{회사명}.xlsx'.
            modules: 포함할 모듈 목록. None이면 데이터가 있는 모든 모듈.
                    finance: "IS", "BS", "CF" (연도별 시계열 피벗)
                    비율: "ratios"
                    나머지: Company property 이름 (dividend, audit, employee, executive, ...)

    Returns:
            저장된 파일 경로 (str).
    """
    allModules = _getAvailableModules(c)
    labelMap = {name: label for name, label in allModules}

    if modules is None:
        targetModules = [name for name, _ in allModules]
    else:
        targetModules = list(dict.fromkeys(modules))

    wb = Workbook()
    wb.remove(wb.active)

    financeModules = [m for m in targetModules if m in _FINANCE_SHEETS]
    if financeModules and _hasFinance(c):
        result = _buildAnnualSeries(c)
        if result:
            series, years = result
            for sjDiv in financeModules:
                if sjDiv in series:
                    _writeFinanceSheet(wb, sjDiv, series, years)

    if "ratios" in targetModules and _hasFinance(c):
        _writeRatiosSheet(wb, c)

    for moduleName in targetModules:
        if moduleName in _FINANCE_SHEETS or moduleName in _SPECIAL_SHEETS:
            continue
        df = _resolveData(c, moduleName)
        if df is not None:
            label = labelMap.get(moduleName, moduleName)
            _writeDataFrameSheet(wb, label, df)

    if not wb.sheetnames:
        raise ValueError(f"{c.stockCode} ({c.corpName}): 내보낼 데이터 없음")

    if outputPath is None:
        safeName = c.corpName.replace("/", "_").replace("\\", "_")
        outputPath = Path.cwd() / f"{c.stockCode}_{safeName}.xlsx"
    else:
        outputPath = Path(outputPath)

    wb.save(str(outputPath))
    return str(outputPath)


def exportWithTemplate(
    c: Company,
    template: ExcelTemplate,
    outputPath: str | Path | None = None,
) -> str:
    """템플릿 기반 Excel 내보내기.

    SheetSpec별로 소스를 로드하고, columns/years 필터를 적용하여 시트를 생성한다.

    Args:
            c: Company 인스턴스.
            template: ExcelTemplate (시트 구성 정의).
            outputPath: 저장 경로. None이면 자동 생성.

    Returns:
            저장된 파일 경로 (str).
    """
    from dartlab.viz.export.template import PanelTableSource

    wb = Workbook()
    wb.remove(wb.active)

    annualCache: tuple[dict, list[str]] | None = None

    for spec in template.sheets:
        source = spec.source

        # 공시 panel 표 — 병합 보존 격자 (회귀 0: ModuleSource 경로는 아래 그대로).
        if isinstance(source, PanelTableSource):
            _writePanelTableSheet(wb, source, c, label=spec.label)
            continue

        # ModuleSource (또는 하위호환 정규화된 문자열 source) — 모듈명으로 분기.
        name = source.name

        if name in _FINANCE_SHEETS:
            if not _hasFinance(c):
                continue
            if annualCache is None:
                result = _buildAnnualSeries(c)
                if result is None:
                    continue
                annualCache = result
            series, years = annualCache
            if name in series:
                _writeFinanceSheet(
                    wb,
                    name,
                    series,
                    years,
                    label=spec.label,
                    columns=spec.columns or None,
                    yearFilter=spec.years or None,
                )
            continue

        if name == "ratios":
            if _hasFinance(c):
                _writeRatiosSheet(wb, c, label=spec.label)
            continue

        df = _resolveData(c, name)
        if df is not None:
            _writeDataFrameSheet(
                wb,
                spec.label,
                df,
                columns=spec.columns or None,
                maxRows=spec.maxRows,
            )

    if not wb.sheetnames:
        raise ValueError(f"{c.stockCode} ({c.corpName}): 템플릿 '{template.name}'에 해당하는 데이터 없음")

    if outputPath is None:
        safeName = c.corpName.replace("/", "_").replace("\\", "_")
        templateSafe = template.name.replace("/", "_").replace("\\", "_")
        outputPath = Path.cwd() / f"{c.stockCode}_{safeName}_{templateSafe}.xlsx"
    else:
        outputPath = Path(outputPath)

    wb.save(str(outputPath))
    return str(outputPath)


def listAvailableModules(c: Company) -> list[dict[str, str]]:
    """내보내기 가능한 모듈 목록 반환 (GUI/API용).

    registry의 requires 필드로 빠르게 판정 (property 호출 없음).
    """
    from dartlab.core.registry import getEntry

    _DATA_FLAGS = {
        "panel": c._hasPanel,
        "finance": _hasFinance(c),
        "report": c._hasReport,
    }

    result = []
    for name, label in _getAvailableModules(c):
        entry = getEntry(name)
        if entry is not None and entry.requires:
            if not _DATA_FLAGS.get(entry.requires, False):
                continue
        elif name in _FINANCE_SHEETS or name == "ratios":
            if not _hasFinance(c):
                continue
        result.append({"name": name, "label": label})
    return result
