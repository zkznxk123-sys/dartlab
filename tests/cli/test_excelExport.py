"""Excel 내보내기 테스트."""

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import polars as pl
import pytest

from dartlab.viz.export.excel import exportToExcel, listAvailableModules

pytestmark = pytest.mark.integration

PATCH_TARGET = "dartlab.providers.dart.finance.pivot.buildAnnual"


class _ShowableMock(MagicMock):
    """show() 지원 MagicMock."""

    _show_data: dict

    def show(self, topic, **kwargs):
        return self._show_data.get(topic)


@pytest.fixture
def mockCompany():
    c = _ShowableMock()
    c._show_data = {}
    c.stockCode = "005930"
    c.corpName = "삼성전자"
    c._hasFinance = True
    c._hasDocs = True
    c._hasReport = True
    c._cache = {"_finance_q_CFS": ({"IS": {"sales": [1, 2, 3]}, "BS": {}, "CF": {}}, ["2022_Q1"])}
    c.report = MagicMock()
    c._report.dividend = None
    c._report.employee = None
    return c


def test_exportToExcel_creates_file(mockCompany):
    series = {
        "IS": {"sales": [100, 200, 300], "net_profit": [10, 20, 30]},
        "BS": {"total_assets": [1000, 2000, 3000]},
        "CF": {"operating_cashflow": [50, 60, 70]},
    }
    years = ["2022", "2023", "2024"]

    with patch(PATCH_TARGET, return_value=(series, years)):
        with tempfile.TemporaryDirectory() as tmpDir:
            outPath = Path(tmpDir) / "test.xlsx"
            result = exportToExcel(mockCompany, outputPath=outPath, modules=["IS", "BS", "CF"])
            assert Path(result).exists()
            assert result.endswith(".xlsx")

            from openpyxl import load_workbook

            wb = load_workbook(result)
            assert "손익계산서" in wb.sheetnames
            assert "재무상태표" in wb.sheetnames
            assert "현금흐름표" in wb.sheetnames


def test_exportToExcel_module_selection(mockCompany):
    series = {
        "IS": {"sales": [100, 200]},
        "BS": {"total_assets": [1000, 2000]},
        "CF": {},
    }
    years = ["2023", "2024"]

    with patch(PATCH_TARGET, return_value=(series, years)):
        with tempfile.TemporaryDirectory() as tmpDir:
            outPath = Path(tmpDir) / "test.xlsx"
            result = exportToExcel(mockCompany, outputPath=outPath, modules=["IS"])

            from openpyxl import load_workbook

            wb = load_workbook(result)
            assert "손익계산서" in wb.sheetnames
            assert "재무상태표" not in wb.sheetnames


def test_exportToExcel_default_filename(mockCompany):
    series = {"IS": {"sales": [100]}, "BS": {}, "CF": {}}
    years = ["2024"]

    with patch(PATCH_TARGET, return_value=(series, years)):
        with tempfile.TemporaryDirectory() as tmpDir:
            import os

            origCwd = os.getcwd()
            os.chdir(tmpDir)
            try:
                result = exportToExcel(mockCompany, modules=["IS"])
                assert "005930" in result
                assert "삼성전자" in result
                assert result.endswith(".xlsx")
            finally:
                os.chdir(origCwd)


def test_exportToExcel_no_data_raises(mockCompany):
    mockCompany._hasFinance = False
    mockCompany._hasReport = False
    mockCompany._hasDocs = False
    mockCompany.report = None

    with pytest.raises(ValueError, match="내보낼 데이터 없음"):
        with tempfile.TemporaryDirectory() as tmpDir:
            exportToExcel(mockCompany, outputPath=Path(tmpDir) / "t.xlsx", modules=["IS"])


def test_ratios_sheet(mockCompany):
    annualSeries = {
        "IS": {"sales": [100, 200, 300], "operating_profit": [10, 20, 30], "net_profit": [5, 10, 15]},
        "BS": {
            "total_assets": [1000, 2000, 3000],
            "owners_of_parent_equity": [500, 600, 700],
            "total_liabilities": [500, 1400, 2300],
        },
        "CF": {"operating_cashflow": [50, 60, 70]},
    }
    years = ["2022", "2023", "2024"]
    mockCompany._buildFinanceSeries = lambda **kwargs: (annualSeries, years)

    with patch(PATCH_TARGET, return_value=(annualSeries, years)):
        with tempfile.TemporaryDirectory() as tmpDir:
            outPath = Path(tmpDir) / "test.xlsx"
            result = exportToExcel(mockCompany, outputPath=outPath, modules=["ratios"])

            from openpyxl import load_workbook

            wb = load_workbook(result)
            assert "재무비율" in wb.sheetnames
            ws = wb["재무비율"]
            assert ws.cell(2, 1).value == "수익성"
            assert ws.cell(2, 2).value is None
            assert ws.cell(3, 1).value is not None
            headers = [ws.cell(1, c).value for c in range(2, 5)]
            assert headers == ["2022", "2023", "2024"]


def test_dataframe_property_as_sheet(mockCompany):
    df = pl.DataFrame({"year": [2022, 2023], "total": [100, 200]})
    mockCompany.dividend = df
    mockCompany._show_data["dividend"] = df
    mockCompany._hasFinance = False

    with tempfile.TemporaryDirectory() as tmpDir:
        outPath = Path(tmpDir) / "test.xlsx"
        result = exportToExcel(mockCompany, outputPath=outPath, modules=["dividend"])

        from openpyxl import load_workbook

        wb = load_workbook(result)
        assert "배당" in wb.sheetnames
        ws = wb["배당"]
        assert ws.cell(1, 1).value == "year"
        assert ws.cell(2, 1).value == 2022


def test_listAvailableModules(mockCompany):
    df = pl.DataFrame({"year": [2022], "val": [1]})
    mockCompany.dividend = df
    mockCompany.audit = None

    result = listAvailableModules(mockCompany)
    names = [m["name"] for m in result]
    assert "IS" in names
    assert "ratios" in names
    assert "dividend" in names
