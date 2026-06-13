"""table-export Phase-1 회귀 가드 — DART XML 정규화·정합 정규화·격자 패리티·source union.

`tests/_attempts/tableExport/` 개념검증을 본진 보강한 회귀 게이트. 순수 unit (coerceCell·
normalizeDartXml·source round-trip·격자 골든 패리티) + realData 1건 (실회사 PanelTableSource
→ .xlsx, openpyxl read-back 으로 merge_cells 비어있지 않음 + Number + 음수 증명).

골든 픽스처 = `tests/fixtures/xmlTables/*.xml` (DART 대문자 원본) + `*.grid.json`
(cellGrid 결과). 엔진·브라우저 패리티 SSOT.

OOM 규율: 직접 pytest 금지 — `bash tests/test-lock.sh tests/parse/test_dartXmlNormalize.py -m "unit" -v`.
realData 는 회사 1개만 (`bash tests/test-lock.sh ... -m "realData" -v`).
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from dartlab.providers.dart.parse.dartXmlNormalize import (
    coerceCell,
    detectUnit,
    normalizeDartXml,
)
from dartlab.providers.dart.parse.htmlTableParser import cellGrid

_FIXTURES = Path(__file__).resolve().parents[1] / "fixtures" / "xmlTables"
_GOLDEN = [
    "small_merged",
    "samsung_3_타법인출자_현황_상세_0",
    "samsung_IX_계열회사_등에_관한_사항_1",
]


# ── coerceCell ──


@pytest.mark.unit
@pytest.mark.parametrize(
    "text,expected",
    [
        ("1,234", 1234),
        ("445,244", 445244),
        ("23.7", 23.7),
        ("(1,234)", -1234),
        ("(2,554,690)", -2554690),
        ("△5", -5),
        ("△1,234.5", -1234.5),
        ("▲500", -500),
        ("△\n1,202,857", -1202857),  # 줄바꿈 낀 삼각형 음수 (실데이터)
        ("삼성전자", "삼성전자"),
        ("", None),
        ("   ", None),
        ("5,000원", "5,000원"),  # 단위 접미 → 문자열 유지 (정직)
        ("2024.12.31", "2024.12.31"),  # 날짜 (점 2개) → 문자열 유지
        ("-", "-"),
        ("△", "△"),
        ("-99", -99),
    ],
)
def test_coerceCell(text, expected):
    assert coerceCell(text) == expected


@pytest.mark.unit
def test_coerceCell_none_never_zero():
    # honest-gap: 결손은 None, 절대 0 아님.
    assert coerceCell("") is None
    assert coerceCell("   ") is None
    assert coerceCell("") != 0


@pytest.mark.unit
def test_coerceCell_types():
    assert isinstance(coerceCell("1,234"), int)
    assert isinstance(coerceCell("23.7"), float)
    assert isinstance(coerceCell("삼성"), str)


# ── normalizeDartXml ──


@pytest.mark.unit
def test_normalizeDartXml_tagMap():
    out = normalizeDartXml('<TABLE><TR><TU COLSPAN="2" ACODE="x">H</TU></TR></TABLE>')
    assert out == '<table><tr><th colspan="2">H</th></tr></table>'


@pytest.mark.unit
def test_normalizeDartXml_TE_to_td():
    out = normalizeDartXml("<TR><TE>v</TE></TR>")
    assert out == "<tr><td>v</td></tr>"


@pytest.mark.unit
def test_normalizeDartXml_dropsGovMeta_keepsStructure():
    out = normalizeDartXml('<TE ROWSPAN="3" ALIGN="right" ACONTEXT="gov" USERMARK="z">v</TE>')
    assert "acontext" not in out.lower()
    assert "usermark" not in out.lower()
    assert 'rowspan="3"' in out
    assert 'align="right"' in out


@pytest.mark.unit
def test_normalizeDartXml_passthrough_noTags():
    assert normalizeDartXml("plain text") == "plain text"
    assert normalizeDartXml("") == ""


# ── detectUnit ──


@pytest.mark.unit
@pytest.mark.parametrize(
    "caption,expected",
    [
        ("(단위: 백만원)", "백만원"),
        ("단위 : 천원", "천원"),
        ("(단위:원)", "원"),
        ("단위 : 원, 주", "원"),
        ("매출 추이", ""),
        ("(단위: 광년)", ""),  # 미지 단위 → 빈칸 (추측 0)
    ],
)
def test_detectUnit(caption, expected):
    assert detectUnit(caption) == expected


# ── 격자 골든 패리티 (엔진 cellGrid vs 브라우저 SSOT *.grid.json) ──


@pytest.mark.unit
@pytest.mark.parametrize("name", _GOLDEN)
def test_grid_parity_against_golden(name):
    xml = (_FIXTURES / f"{name}.xml").read_text(encoding="utf-8")
    golden = json.loads((_FIXTURES / f"{name}.grid.json").read_text(encoding="utf-8"))
    grid = cellGrid(normalizeDartXml(xml))

    # shape 일치
    assert len(grid) == golden["shape"]["rows"], f"{name}: row count"
    nCols = max((len(r) for r in grid), default=0)
    assert nCols == golden["shape"]["cols"], f"{name}: col count"

    # 셀별 text/colspan/rowspan/align/isHeader 일치 (병합셀 = 같은 인스턴스 → 같은 dict)
    g = golden["grid"]
    for r, row in enumerate(grid):
        for cIdx, cell in enumerate(row):
            exp = g[r][cIdx]
            if cell is None:
                assert exp is None, f"{name}: ({r},{cIdx}) None mismatch"
                continue
            assert cell.text == exp["text"], f"{name}: ({r},{cIdx}) text"
            assert cell.colspan == exp["colspan"], f"{name}: ({r},{cIdx}) colspan"
            assert cell.rowspan == exp["rowspan"], f"{name}: ({r},{cIdx}) rowspan"
            assert cell.align == exp["align"], f"{name}: ({r},{cIdx}) align"
            assert cell.isHeader == exp["isHeader"], f"{name}: ({r},{cIdx}) isHeader"


@pytest.mark.unit
def test_grid_merge_shared_instance():
    # small_merged: 제 46 기 colspan=2 → 같은 인스턴스가 (0,0),(0,1).
    xml = (_FIXTURES / "small_merged.xml").read_text(encoding="utf-8")
    grid = cellGrid(normalizeDartXml(xml))
    assert grid[0][0] is grid[0][1]
    assert grid[0][0].colspan == 2


# ── source union round-trip + 하위호환 ──


@pytest.mark.unit
def test_source_str_becomes_moduleSource():
    from dartlab.viz.export.template import ModuleSource, SheetSpec

    s = SheetSpec(source="IS")
    assert isinstance(s.source, ModuleSource)
    assert s.source.name == "IS"
    assert s.label == "IS"  # 빈 라벨 → name 파생


@pytest.mark.unit
def test_source_dict_branches_to_panelTable():
    from dartlab.viz.export.template import PanelTableSource, SheetSpec

    s = SheetSpec(
        source={
            "kind": "panelTable",
            "chapter": "I. 회사의 개요",
            "sectionLeaf": "1. 회사의 개요",
            "leafType": "table",
            "leafSeq": 3,
            "periodMode": "asFiled",
            "period": "2025Q4",
        }
    )
    assert isinstance(s.source, PanelTableSource)
    assert s.source.chapter == "I. 회사의 개요"
    assert s.source.leafSeq == 3
    assert s.source.period == "2025Q4"


@pytest.mark.unit
def test_presets_roundtrip_unchanged():
    from dartlab.viz.export.template import PRESETS, ExcelTemplate, ModuleSource

    for key, t in PRESETS.items():
        back = ExcelTemplate.fromJson(t.toJson())
        assert len(back.sheets) == len(t.sheets), key
        for orig, rt in zip(t.sheets, back.sheets):
            assert isinstance(rt.source, ModuleSource), key
            assert rt.source.name == orig.source.name, key
            assert rt.label == orig.label, key


@pytest.mark.unit
def test_saved_string_source_template_loads_unchanged():
    from dartlab.viz.export.template import ExcelTemplate

    saved = '{"name": "내양식", "sheets": [{"source": "IS", "label": "손익"}, {"source": "dividend"}]}'
    t = ExcelTemplate.fromJson(saved)
    assert t.sheets[0].source.name == "IS"
    assert t.sheets[0].label == "손익"
    assert t.sheets[1].source.name == "dividend"
    assert t.sheets[1].label == "dividend"


@pytest.mark.unit
def test_panelTable_source_roundtrip():
    from dartlab.viz.export.template import (
        ExcelTemplate,
        PanelTableSource,
        SheetSpec,
    )

    t = ExcelTemplate(
        name="t",
        sheets=[
            SheetSpec(
                source=PanelTableSource(
                    kind="panelTable",
                    chapter="III. 재무에 관한 사항",
                    sectionLeaf="2. 연결재무제표",
                    disclosureKey="BS_C",
                    scope="consolidated",
                    leafSeq=0,
                    periodMode="horizontalized",
                ),
                label="연결BS",
            )
        ],
    )
    back = ExcelTemplate.fromJson(t.toJson())
    src = back.sheets[0].source
    assert isinstance(src, PanelTableSource)
    assert src.disclosureKey == "BS_C"
    assert src.scope == "consolidated"
    assert src.periodMode == "horizontalized"
    assert src.leafSeq == 0


# ── realData: 실회사 PanelTableSource → .xlsx (merge + Number + 음수 증명) ──


@pytest.mark.realData
def test_real_company_panelTable_export(tmp_path):
    """005930 의 병합 보유 공시 표 1개를 .xlsx 로 — openpyxl read-back 으로 병합·Number·음수 증명."""
    import openpyxl
    import polars as pl
    from openpyxl.cell.cell import MergedCell

    from dartlab.providers.dart.panel.read import readWide
    from dartlab.viz.export.excel import _selectPanelRow, _writePanelTableSheet
    from dartlab.viz.export.template import PanelTableSource

    wide = readWide("005930", marketNs="kr", tag=True)
    assert wide is not None, "panel artifact 부재 — realData skip 환경"

    # 병합(colspan/rowspan)을 가진 table 행 + Number/음수 가능성 높은 표를 하나 찾는다.
    idxCols = {"chapter", "sectionLeaf", "blockLeaf", "leafType", "disclosureKey", "scope", "leafSeq"}
    periodCols = [c for c in wide.columns if c not in idxCols]
    tblRows = wide.filter(pl.col("leafType") == "table")

    chosen = None
    for row in tblRows.iter_rows(named=True):
        for pc in periodCols:
            v = row.get(pc)
            if v and "<TABLE" in v.upper() and ("COLSPAN=" in v.upper() or "ROWSPAN=" in v.upper()):
                chosen = (row, pc)
                break
        if chosen:
            break
    assert chosen is not None, "병합 보유 table 행을 못 찾음"
    row, period = chosen

    spec = PanelTableSource(
        kind="panelTable",
        chapter=row.get("chapter") or "",
        sectionLeaf=row.get("sectionLeaf") or "",
        blockLeaf=row.get("blockLeaf") or "",
        leafType="table",
        disclosureKey=row.get("disclosureKey"),
        scope=row.get("scope"),
        leafSeq=row.get("leafSeq"),
        periodMode="asFiled",
        period=period,
    )
    # _selectPanelRow 가 같은 행을 다시 찾는지 확인 (식별 계약).
    picked = _selectPanelRow(wide, spec)
    assert picked is not None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    class _StubCompany:
        # _writePanelTableSheet 는 _loadPanelWide(c) → readWide(c.stockCode, marketNs) 로
        # 같은 회사를 로드하므로 stockCode/marketNs 만 있으면 실제 export 경로를 그대로 탄다.
        stockCode = "005930"
        marketNs = "kr"

    _writePanelTableSheet(wb, spec, _StubCompany(), label="공시표")
    assert wb.sheetnames, "시트가 생성되지 않음"

    out = tmp_path / "real_export.xlsx"
    wb.save(str(out))

    # openpyxl read-back
    rb = openpyxl.load_workbook(str(out))
    ws = rb[rb.sheetnames[0]]

    # 1) merge_cells 비어있지 않음 (병합 보존)
    assert len(ws.merged_cells.ranges) > 0, "merge_cells 비어있음 — 병합 보존 실패"

    # 2) Number 셀 존재 + (가능하면) 음수
    nums = []
    for r in ws.iter_rows():
        for cell in r:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                nums.append(cell.value)
    assert any(nums), "Number 셀이 없음 — coerceCell 배선 실패"

    # 음수는 표마다 유무가 다르므로, 이 표에 없으면 전사 스캔으로 음수 1개 증명 (계약 검증).
    if not any(n < 0 for n in nums):
        neg_found = False
        from dartlab.providers.dart.parse.dartXmlNormalize import coerceCell, normalizeDartXml
        from dartlab.providers.dart.parse.htmlTableParser import cellGrid

        for r2 in tblRows.head(80).iter_rows(named=True):
            for pc in periodCols:
                v = r2.get(pc)
                if not v or "<TABLE" not in v.upper():
                    continue
                for gr in cellGrid(normalizeDartXml(v)):
                    for cl in gr:
                        if cl is None:
                            continue
                        cv = coerceCell(cl.text)
                        if isinstance(cv, (int, float)) and not isinstance(cv, bool) and cv < 0:
                            neg_found = True
                            break
                    if neg_found:
                        break
                if neg_found:
                    break
            if neg_found:
                break
        assert neg_found, "음수 정규화 셀을 전사 스캔에서도 못 찾음"

    del wide, tblRows
